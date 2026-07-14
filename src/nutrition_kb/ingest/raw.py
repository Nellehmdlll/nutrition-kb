#!/usr/bin/env python3
"""Ingestion raw : MANIFEST.json (SHA-256) + extraction xlsx -> Parquet par feuille."""

import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq

SCRIPT_VERSION = "3.0.0"

ROOT_DIR = Path(__file__).resolve().parents[3]
SOURCES_DIR = ROOT_DIR / "data" / "raw" / "sources"
EXTRACTED_DIR = ROOT_DIR / "data" / "raw" / "extracted"
METADATA_FILE = SOURCES_DIR / "sources_metadata.json"
MANIFEST_FILE = SOURCES_DIR / "MANIFEST.json"
CHUNK_SIZE = 1024 * 1024


class SourceChangedError(RuntimeError):
    """Le hash d'un fichier source a change depuis la derniere ingestion."""


class FormulaFoundError(RuntimeError):
    """Une formule Excel a ete trouvee dans le classeur source."""


def sha256_of(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(CHUNK_SIZE), b""):
            digest.update(chunk)
    return digest.hexdigest()


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


# --- MANIFEST.json (hash de chaque fichier source) ---------------------------


def generate_manifest() -> list[dict]:
    metadata = (
        json.loads(METADATA_FILE.read_text(encoding="utf-8"))
        if METADATA_FILE.exists()
        else {}
    )
    previous = {}
    if MANIFEST_FILE.exists():
        for entry in json.loads(MANIFEST_FILE.read_text(encoding="utf-8")):
            previous[entry["name"]] = entry

    skip_names = {METADATA_FILE.name, MANIFEST_FILE.name}
    entries = []
    missing_metadata = []

    for path in sorted(SOURCES_DIR.iterdir()):
        if not path.is_file() or path.name in skip_names:
            continue

        current_hash = sha256_of(path)
        prev_entry = previous.get(path.name)
        if prev_entry is not None and prev_entry["sha256"] != current_hash:
            raise SourceChangedError(
                f"{path.name} : le SHA-256 a change depuis le dernier MANIFEST.json "
                f"({prev_entry['sha256']} -> {current_hash}). "
                "Archivez/renommez l'ancienne version avant de reingerer."
            )
        ingestion_date = prev_entry["ingestion_date_utc"] if prev_entry else utc_now_iso()

        entry_meta = metadata.get(path.name, {})
        url = entry_meta.get("url")
        license_ = entry_meta.get("license")
        if not url or not license_:
            missing_metadata.append(path.name)

        entries.append(
            {
                "name": path.name,
                "sha256": current_hash,
                "size_bytes": path.stat().st_size,
                "ingestion_date_utc": ingestion_date,
                "source_url": url,
                "license": license_,
                "script_version": SCRIPT_VERSION,
            }
        )

    MANIFEST_FILE.write_text(
        json.dumps(entries, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    if missing_metadata:
        print(
            f"[manifest] url/license manquants pour : {', '.join(missing_metadata)}",
            file=sys.stderr,
        )
    return entries


# --- Extraction xlsx -> Parquet (une feuille = un fichier) --------------------


def slugify(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", name.strip().lower())
    return slug.strip("_")


def read_sheet_cells(ws) -> list[tuple[int, list]]:
    """Lit toute la feuille puis la borne a la vraie boite englobante des
    valeurs non vides (ignore le formatage cosmetique hors-donnees)."""
    all_rows = [
        (row_num, [cell.value for cell in row])
        for row_num, row in enumerate(ws.iter_rows(), start=1)
    ]

    max_row = 0
    max_col = 0
    for row_num, values in all_rows:
        for col_idx, v in enumerate(values, start=1):
            if v is not None:
                max_row = max(max_row, row_num)
                max_col = max(max_col, col_idx)

    if max_row == 0:
        return []
    return [(row_num, values[:max_col]) for row_num, values in all_rows if row_num <= max_row]


def build_sheet_dataframe(
    ws, source_file_name: str, sheet_name: str, ingested_at: str
) -> pd.DataFrame:
    rows = read_sheet_cells(ws)
    n_cols = max((len(values) for _, values in rows), default=0)
    col_names = [f"col_{i}" for i in range(n_cols)]

    records = []
    for row_num, values in rows:
        record = {
            col_names[i]: (None if v is None else str(v)) for i, v in enumerate(values)
        }
        record["_source_file"] = source_file_name
        record["_source_sheet"] = sheet_name
        record["_source_row"] = row_num
        record["_ingested_at"] = ingested_at
        records.append(record)

    columns = col_names + ["_source_file", "_source_sheet", "_source_row", "_ingested_at"]
    return pd.DataFrame.from_records(records, columns=columns)


def build_sheet_schema(n_data_cols: int) -> pa.Schema:
    """Type explicitement chaque colonne : une colonne entierement vide ne
    doit jamais etre devinee comme numerique (NaN) par pyarrow."""
    fields = [pa.field(f"col_{i}", pa.string()) for i in range(n_data_cols)]
    fields += [
        pa.field("_source_file", pa.string()),
        pa.field("_source_sheet", pa.string()),
        pa.field("_source_row", pa.int64()),
        pa.field("_ingested_at", pa.string()),
    ]
    return pa.schema(fields)


def write_parquet_with_metadata(
    df: pd.DataFrame, out_path: Path, schema: pa.Schema, extra_metadata: dict
) -> None:
    table = pa.Table.from_pandas(df, schema=schema, preserve_index=False)
    meta = dict(table.schema.metadata or {})
    for key, value in extra_metadata.items():
        meta[f"nutrition_kb.{key}".encode()] = str(value).encode()
    table = table.replace_schema_metadata(meta)
    pq.write_table(table, out_path)


def assert_no_formulas(xlsx_path: Path) -> None:
    """data_only=True rend None pour une formule sans valeur en cache : un trou
    silencieux dans les donnees. On refuse toute formule plutot que de risquer ca."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
    offenders = []
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    offenders.append(f"{sheet_name}!{cell.coordinate}")
    if offenders:
        raise FormulaFoundError(
            f"{len(offenders)} cellule(s) contiennent une formule dans {xlsx_path.name} : "
            f"{', '.join(offenders[:20])}"
            + (" ..." if len(offenders) > 20 else "")
            + ". Ingestion refusee : une formule sans valeur en cache serait lue comme None."
        )


def extract_workbook(xlsx_path: Path, output_dir: Path) -> dict:
    assert_no_formulas(xlsx_path)

    source_sha256 = sha256_of(xlsx_path)
    index_file = output_dir / "sheet_index.json"

    if index_file.exists():
        previous = json.loads(index_file.read_text(encoding="utf-8"))
        if previous["source_sha256"] != source_sha256:
            raise SourceChangedError(
                f"{xlsx_path.name} : le SHA-256 a change depuis la derniere extraction "
                f"({previous['source_sha256']} -> {source_sha256}). "
                "Archivez/renommez l'ancien dossier extrait avant de reingerer."
            )
        ingested_at = previous["ingested_at_utc"]
    else:
        ingested_at = utc_now_iso()

    output_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    used_slugs: dict[str, int] = {}
    sheets_meta = []

    for position, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        df = build_sheet_dataframe(ws, xlsx_path.name, sheet_name, ingested_at)

        slug = slugify(sheet_name)
        used_slugs[slug] = used_slugs.get(slug, 0) + 1
        if used_slugs[slug] > 1:
            slug = f"{slug}_{used_slugs[slug]}"
        out_path = output_dir / f"{slug}.parquet"

        n_data_cols = sum(1 for c in df.columns if not str(c).startswith("_"))
        schema = build_sheet_schema(n_data_cols)
        write_parquet_with_metadata(
            df,
            out_path,
            schema,
            {
                "source_file": xlsx_path.name,
                "source_sha256": source_sha256,
                "source_sheet": sheet_name,
                "script_version": SCRIPT_VERSION,
                "ingested_at_utc": ingested_at,
            },
        )

        sheets_meta.append(
            {
                "sheet_position": position,
                "sheet_name": sheet_name,
                "parquet_file": out_path.name,
                "parquet_sha256": sha256_of(out_path),
                "rows": df.shape[0],
                "cols": n_data_cols,
            }
        )
        print(f"[extract] [{position:2d}] {sheet_name!r:45} -> {out_path.name}  ({df.shape[0]} lignes)")

    index_payload = {
        "source_file": xlsx_path.name,
        "source_sha256": source_sha256,
        "ingested_at_utc": ingested_at,
        "script_version": SCRIPT_VERSION,
        "sheets": sheets_meta,
    }
    index_file.write_text(
        json.dumps(index_payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    return index_payload


# --- CLI -----------------------------------------------------------------------


def main() -> int:
    try:
        generate_manifest()
    except SourceChangedError as e:
        print(f"[manifest] INGESTION BLOQUEE : {e}", file=sys.stderr)
        return 2

    xlsx_path = SOURCES_DIR / "WAFCT_2019.xlsx"
    if not xlsx_path.exists():
        print(f"Fichier introuvable : {xlsx_path}", file=sys.stderr)
        return 1

    try:
        extract_workbook(xlsx_path, EXTRACTED_DIR / "wafct_2019")
    except SourceChangedError as e:
        print(f"[extract] INGESTION BLOQUEE : {e}", file=sys.stderr)
        return 2
    except FormulaFoundError as e:
        print(f"[extract] INGESTION BLOQUEE : {e}", file=sys.stderr)
        return 3

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
